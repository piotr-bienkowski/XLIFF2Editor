"""
XLIFF 2.2 → Excel merger.

Reads translations from an XLIFF 2.2 file produced by excel_xliff22_converter
and writes them back to the original Excel workbook in-place (target column).
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

NS22 = 'urn:oasis:names:tc:xliff:document:2.0'


def _has_trailing_ph(elem) -> bool:
    """True if element's last child is a <ph> with empty/whitespace tail.

    The converter marks segment boundaries by appending a trailing <ph> with
    an empty tail.  Translation tools may change the equiv value (e.g. \\n →
    space), so we detect by position rather than by equiv value.
    """
    if elem is None:
        return False
    children = list(elem)
    if not children:
        return False
    return not (children[-1].tail or '').strip()


def _target_text(tgt_elem) -> str:
    """Read target text, handling inline <ph> tags.

    - Trailing <ph> tags (empty/whitespace tail) are stripped — they are
      segment-boundary markers and must not appear in the merged cell value.
    - Inner <ph equiv="\\n"/> tags are converted to \\n.
    - Other inner <ph> tags use their equiv value as-is.
    """
    if tgt_elem is None:
        return ''
    children = list(tgt_elem)
    # Drop trailing boundary ph tags
    while children and not (children[-1].tail or '').strip():
        children = children[:-1]
    parts = [tgt_elem.text or '']
    for child in children:
        equiv = child.get('equiv', '')
        parts.append('\n' if equiv == '\n' else equiv)
        parts.append(child.tail or '')
    return ''.join(parts)


def merge_xliff22_to_excel(xliff_path, excel_path) -> dict:
    """
    Write translations from *xliff_path* into the target column of *excel_path*.

    The XLIFF must have been produced by convert_excel_to_xliff22 (requires the
    ``x-excel-tgt-col`` attribute on the ``<file>`` element).

    Multiple units sharing the same ``x-excel-row`` are joined with a space,
    except where the source segment ends with a trailing ``<ph>`` boundary
    marker — those boundaries are rejoined with ``\\n``.

    Args:
        xliff_path:  Path to the translated XLIFF 2.2 file.
        excel_path:  Path to the Excel workbook to update in-place.

    Returns:
        {'rows_written': int, 'units_skipped': int}

    Raises:
        ValueError: if the XLIFF was not created by the Excel converter.
    """
    xliff_path = Path(xliff_path)
    excel_path = Path(excel_path)

    tree = etree.parse(str(xliff_path))
    root = tree.getroot()

    file_elem = root.find(f'{{{NS22}}}file')
    if file_elem is None:
        raise ValueError("No <file> element found in XLIFF.")
    tgt_col = file_elem.get('x-excel-tgt-col')
    if not tgt_col:
        raise ValueError(
            "This file was not imported from Excel (missing x-excel-tgt-col)."
        )
    tgt_col_idx = column_index_from_string(tgt_col)

    # Group (text, trailing_ph) pairs by Excel row (preserve document order).
    # trailing_ph=True → use \n as separator before the next segment.
    row_items: dict[int, list[tuple[str, bool]]] = defaultdict(list)
    units_skipped = 0

    for unit in root.findall(f'.//{{{NS22}}}unit'):
        row_str = unit.get('x-excel-row')
        if not row_str:
            units_skipped += 1
            continue
        row = int(row_str)
        for seg in unit.findall(f'{{{NS22}}}segment'):
            tgt = seg.find(f'{{{NS22}}}target')
            src = seg.find(f'{{{NS22}}}source')
            text = _target_text(tgt).strip()
            if text:
                row_items[row].append((text, _has_trailing_ph(src)))

    wb = load_workbook(str(excel_path))
    ws = wb.active
    max_row = ws.max_row or 0
    rows_written = 0

    for row, items in row_items.items():
        parts: list[str] = []
        for i, (text, trailing_ph) in enumerate(items):
            parts.append(text)
            if i < len(items) - 1:
                parts.append('\n' if trailing_ph else ' ')
        joined = ''.join(parts)
        if not joined:
            units_skipped += 1
            continue
        if row > max_row:
            units_skipped += 1
            continue
        ws.cell(row=row, column=tgt_col_idx).value = joined
        rows_written += 1

    wb.save(str(excel_path))
    return {'rows_written': rows_written, 'units_skipped': units_skipped}
