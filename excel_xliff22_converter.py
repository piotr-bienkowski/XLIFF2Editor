"""
Excel bilingual → XLIFF 2.2 converter.

Standalone converter function + PyQt6 import dialog.
"""

from __future__ import annotations

import re
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from PyQt6.QtWidgets import (
    QDialog, QDialogButtonBox, QFormLayout,
    QLabel, QLineEdit, QSpinBox, QCheckBox,
)

NS22 = 'urn:oasis:names:tc:xliff:document:2.0'
_COL_RE   = re.compile(r'^[A-Za-z]{1,3}$')
_NL_SPLIT = re.compile(r'\r?\n(?=[A-Z])')   # \n before uppercase → segment boundary
_NL_TAG   = re.compile(r'\r?\n')             # any remaining \n → <ph> tag


def _build_source(ns: str, text: str, trailing_nl: bool) -> etree._Element:
    """Build a <source> element, converting \\n to <ph equiv="\\n"/> inline codes."""
    src = etree.Element(f'{{{ns}}}source')
    parts = _NL_TAG.split(text)
    src.text = parts[0]
    ph_id = 1
    for part in parts[1:]:
        ph = etree.SubElement(src, f'{{{ns}}}ph')
        ph.set('id', str(ph_id))
        ph.set('equiv', '\n')
        ph.tail = part
        ph_id += 1
    if trailing_nl:
        ph = etree.SubElement(src, f'{{{ns}}}ph')
        ph.set('id', str(ph_id))
        ph.set('equiv', '\n')
        ph.tail = ''
    return src


# ── import dialog ─────────────────────────────────────────────────────────────

class ExcelImportDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import from Excel")
        layout = QFormLayout(self)

        self.src_lang  = QLineEdit('en-US')
        self.tgt_lang  = QLineEdit('pl-PL')
        self.src_col   = QLineEdit('A')
        self.tgt_col   = QLineEdit('B')
        self.first_row = QSpinBox()
        self.first_row.setMinimum(1)
        self.first_row.setValue(2)
        self.segment_cb = QCheckBox()
        self._col_error = QLabel()
        self._col_error.setStyleSheet('color: red;')

        layout.addRow('Source language:', self.src_lang)
        layout.addRow('Target language:', self.tgt_lang)
        layout.addRow('Source column:',  self.src_col)
        layout.addRow('Target column:',  self.tgt_col)
        layout.addRow('', self._col_error)
        layout.addRow('First data row:', self.first_row)
        layout.addRow('Segment source:', self.segment_cb)

        self._buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        self._buttons.accepted.connect(self.accept)
        self._buttons.rejected.connect(self.reject)
        layout.addRow(self._buttons)

        self.src_col.textChanged.connect(self._validate_cols)
        self.tgt_col.textChanged.connect(self._validate_cols)
        self._validate_cols()

    def _validate_cols(self) -> None:
        ok = bool(_COL_RE.match(self.src_col.text())) and \
             bool(_COL_RE.match(self.tgt_col.text()))
        self._col_error.setText('' if ok else 'Invalid column (use letters only, e.g. A, B, AA)')
        self._buttons.button(QDialogButtonBox.StandardButton.Ok).setEnabled(ok)

    def values(self) -> dict:
        return {
            'src_lang':  self.src_lang.text().strip(),
            'tgt_lang':  self.tgt_lang.text().strip(),
            'src_col':   self.src_col.text().upper().strip(),
            'tgt_col':   self.tgt_col.text().upper().strip(),
            'first_row': self.first_row.value(),
            'segment':   self.segment_cb.isChecked(),
        }


# ── converter ────────────────────────────────────────────────────────────────

def convert_excel_to_xliff22(
    input_path,
    output_path,
    src_lang: str,
    tgt_lang: str,
    src_col: str,
    tgt_col: str,
    first_row: int,
    segment: bool,
    srx_path=None,
) -> dict:
    """
    Convert a bilingual Excel workbook to XLIFF 2.2.

    Args:
        input_path:  Path to the source .xlsx file.
        output_path: Path where the XLIFF will be written.
        src_lang:    BCP-47 source language tag (e.g. 'en-US').
        tgt_lang:    BCP-47 target language tag (e.g. 'pl-PL').
        src_col:     Excel column letter(s) for the source (e.g. 'A').
        tgt_col:     Excel column letter(s) for the target (e.g. 'B').
        first_row:   1-based first data row (rows above are ignored).
        segment:     When True, split source cells into sentence-level units.
        srx_path:    Path to segment.srx; defaults to segment.srx next to this file.

    Returns:
        {'total_units': int, 'total_rows': int}
    """
    input_path  = Path(input_path)
    output_path = Path(output_path)
    src_col_idx = column_index_from_string(src_col.upper())

    segmenter = None
    if segment:
        from srx_segmenter import SrxSegmenter
        segmenter = SrxSegmenter(srx_path)

    wb = load_workbook(str(input_path), read_only=True, data_only=True)
    ws = wb.active

    xliff_root = etree.Element(f'{{{NS22}}}xliff', nsmap={None: NS22})
    xliff_root.set('version', '2.2')
    xliff_root.set('srcLang', src_lang)
    xliff_root.set('trgLang', tgt_lang)

    file_elem = etree.SubElement(xliff_root, f'{{{NS22}}}file')
    file_elem.set('id',              input_path.name)
    file_elem.set('original',        input_path.name)
    file_elem.set('x-excel-src-col', src_col.upper())
    file_elem.set('x-excel-tgt-col', tgt_col.upper())

    unit_counter = 0
    total_rows   = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=first_row, values_only=True), start=first_row
    ):
        value = row[src_col_idx - 1]  # iter_rows tuples are 0-indexed
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        total_rows += 1

        # Build (sentence, trailing_newline_tag) pairs.
        # With segmenter: pre-split on \n+uppercase, then SRX within each part.
        # Without segmenter: one item per row; \n still converted to <ph> tags.
        if segmenter:
            nl_parts = [p for p in _NL_SPLIT.split(text) if p.strip()]
            items = []
            for pi, part in enumerate(nl_parts):
                is_last = (pi == len(nl_parts) - 1)
                sents = segmenter.segment(part.strip(), src_lang)
                for si, sent in enumerate(sents):
                    items.append((sent, si == len(sents) - 1 and not is_last))
        else:
            items = [(text, False)]

        for sentence, trailing_nl in items:
            unit_counter += 1
            unit_elem = etree.SubElement(file_elem, f'{{{NS22}}}unit')
            unit_elem.set('id',          str(unit_counter))
            unit_elem.set('x-excel-row', str(row_idx))
            seg_elem = etree.SubElement(unit_elem, f'{{{NS22}}}segment')
            seg_elem.set('id', str(unit_counter))
            seg_elem.append(_build_source(NS22, sentence, trailing_nl))
            etree.SubElement(seg_elem, f'{{{NS22}}}target')

    wb.close()

    tree = etree.ElementTree(xliff_root)
    tree.write(
        str(output_path),
        xml_declaration=True,
        encoding='UTF-8',
        pretty_print=True,
    )

    return {'total_units': unit_counter, 'total_rows': total_rows}
