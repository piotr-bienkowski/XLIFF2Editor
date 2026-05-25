import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import tempfile
import pytest
from openpyxl import Workbook, load_workbook  # noqa: F401  (load_workbook used in Task 4)
from lxml import etree

NS22 = 'urn:oasis:names:tc:xliff:document:2.0'
SRX_PATH = Path(__file__).parent.parent / 'segment.srx'
_TEMP_FILES: list = []


def teardown_module(_module):
    import os
    for p in _TEMP_FILES:
        try:
            os.unlink(p)
        except OSError:
            pass


# ── helpers ──────────────────────────────────────────────────────────────────

def _make_excel(rows, first_row=2):
    """Create temp xlsx; rows is list of source strings starting at first_row."""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Source'
    ws['B1'] = 'Target'
    for i, text in enumerate(rows, start=first_row):
        ws.cell(row=i, column=1, value=text)
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    _TEMP_FILES.append(tmp.name)
    return Path(tmp.name)


def _make_xliff(units, tgt_col='B'):
    """units = [(row, source_text, target_text), ...]. Returns path to temp xliff."""
    root = etree.Element(f'{{{NS22}}}xliff', nsmap={None: NS22})
    root.set('version', '2.2')
    root.set('srcLang', 'en-US')
    root.set('trgLang', 'pl-PL')
    file_elem = etree.SubElement(root, f'{{{NS22}}}file')
    file_elem.set('id', 'test.xlsx')
    file_elem.set('original', 'test.xlsx')
    file_elem.set('x-excel-src-col', 'A')
    file_elem.set('x-excel-tgt-col', tgt_col)
    for i, (row, src, tgt) in enumerate(units, 1):
        unit = etree.SubElement(file_elem, f'{{{NS22}}}unit')
        unit.set('id', str(i))
        unit.set('x-excel-row', str(row))
        seg = etree.SubElement(unit, f'{{{NS22}}}segment')
        seg.set('id', str(i))
        src_e = etree.SubElement(seg, f'{{{NS22}}}source')
        src_e.text = src
        tgt_e = etree.SubElement(seg, f'{{{NS22}}}target')
        tgt_e.text = tgt if tgt else None
    tree = etree.ElementTree(root)
    tmp = tempfile.NamedTemporaryFile(suffix='.xliff', delete=False)
    tree.write(tmp.name, xml_declaration=True, encoding='UTF-8', pretty_print=True)
    tmp.close()
    _TEMP_FILES.append(tmp.name)
    return Path(tmp.name)


# ── SrxSegmenter tests ───────────────────────────────────────────────────────

from srx_segmenter import SrxSegmenter


@pytest.fixture(scope='module')
def segmenter():
    return SrxSegmenter(SRX_PATH)


def test_segmenter_english_two_sentences(segmenter):
    result = segmenter.segment("Hello world. How are you?", "en-US")
    assert len(result) == 2
    assert result[0] == "Hello world."
    assert result[1] == "How are you?"


def test_segmenter_three_sentences(segmenter):
    result = segmenter.segment("First sentence. Second sentence. Third sentence.", "en-US")
    assert len(result) == 3


def test_segmenter_single_sentence_no_split(segmenter):
    result = segmenter.segment("Hello world", "en-US")
    assert result == ["Hello world"]


def test_segmenter_empty_text(segmenter):
    assert segmenter.segment("", "en-US") == []


def test_segmenter_whitespace_only(segmenter):
    result = segmenter.segment("   ", "en-US")
    assert result == ["   "]


def test_segmenter_polish_abbreviation_no_split(segmenter):
    # "adw." (adwokat) is a Polish abbreviation — should not split
    result = segmenter.segment("Adw. Kowalski złożył wniosek.", "pl-PL")
    assert len(result) == 1


def test_segmenter_returns_list_of_strings(segmenter):
    result = segmenter.segment("Hello. World.", "en-US")
    assert isinstance(result, list)
    assert all(isinstance(r, str) for r in result)


# ── excel_xliff22_converter tests ────────────────────────────────────────────

from excel_xliff22_converter import convert_excel_to_xliff22


def _out_xliff():
    tmp = tempfile.NamedTemporaryFile(suffix='.xliff', delete=False)
    tmp.close()
    _TEMP_FILES.append(tmp.name)
    return Path(tmp.name)


def test_converter_basic():
    excel = _make_excel(['Hello world.', 'How are you?'])
    out = _out_xliff()
    result = convert_excel_to_xliff22(excel, out, 'en-US', 'pl-PL', 'A', 'B', 2, False)
    assert result['total_rows'] == 2
    assert result['total_units'] == 2

    tree = etree.parse(str(out))
    root = tree.getroot()
    units = root.findall(f'.//{{{NS22}}}unit')
    assert len(units) == 2
    assert units[0].get('x-excel-row') == '2'
    assert units[1].get('x-excel-row') == '3'
    sources = root.findall(f'.//{{{NS22}}}source')
    assert sources[0].text == 'Hello world.'
    assert sources[1].text == 'How are you?'
    targets = root.findall(f'.//{{{NS22}}}target')
    assert all(t.text is None for t in targets)


def test_converter_skips_empty_source():
    excel = _make_excel(['Hello.', None, 'World.'])
    out = _out_xliff()
    result = convert_excel_to_xliff22(excel, out, 'en-US', 'pl-PL', 'A', 'B', 2, False)
    assert result['total_rows'] == 2
    assert result['total_units'] == 2


def test_converter_file_metadata():
    excel = _make_excel(['Hello.'])
    out = _out_xliff()
    convert_excel_to_xliff22(excel, out, 'en-US', 'pl-PL', 'A', 'B', 2, False)
    tree = etree.parse(str(out))
    root = tree.getroot()
    assert root.get('srcLang') == 'en-US'
    assert root.get('trgLang') == 'pl-PL'
    file_elem = root.find(f'{{{NS22}}}file')
    assert file_elem.get('x-excel-src-col') == 'A'
    assert file_elem.get('x-excel-tgt-col') == 'B'


def test_converter_segmentation_splits_units():
    excel = _make_excel(['First sentence. Second sentence.'])
    out = _out_xliff()
    result = convert_excel_to_xliff22(
        excel, out, 'en-US', 'pl-PL', 'A', 'B', 2, True,
        srx_path=SRX_PATH
    )
    assert result['total_rows'] == 1
    assert result['total_units'] == 2
    tree = etree.parse(str(out))
    units = tree.getroot().findall(f'.//{{{NS22}}}unit')
    assert len(units) == 2
    # both units refer to the same Excel row
    assert units[0].get('x-excel-row') == '2'
    assert units[1].get('x-excel-row') == '2'


def test_converter_column_b():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws['B2'] = 'Source in B'
    import tempfile as _tmp
    t = _tmp.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(t.name)
    t.close()
    _TEMP_FILES.append(t.name)
    out = _out_xliff()
    result = convert_excel_to_xliff22(t.name, out, 'en-US', 'pl-PL', 'B', 'C', 2, False)
    assert result['total_rows'] == 1
    tree = etree.parse(str(out))
    src = tree.getroot().find(f'.//{{{NS22}}}source')
    assert src.text == 'Source in B'


# ── xliff22_to_excel_merger tests ────────────────────────────────────────────

from xliff22_to_excel_merger import merge_xliff22_to_excel


def _make_excel_for_merge(rows):
    """Create temp xlsx with source rows starting at row 2 (col A only)."""
    wb = Workbook()
    ws = wb.active
    for i, text in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=text)
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    _TEMP_FILES.append(tmp.name)
    return Path(tmp.name)


def test_merger_basic():
    xliff = _make_xliff([
        (2, 'Hello world.', 'Witaj świecie.'),
        (3, 'How are you?', 'Jak się masz?'),
    ])
    excel = _make_excel_for_merge(['Hello world.', 'How are you?'])
    result = merge_xliff22_to_excel(xliff, excel)
    assert result['rows_written'] == 2
    wb = load_workbook(str(excel))
    ws = wb.active
    assert ws.cell(row=2, column=2).value == 'Witaj świecie.'
    assert ws.cell(row=3, column=2).value == 'Jak się masz?'


def test_merger_joins_segmented_units():
    xliff = _make_xliff([
        (2, 'First.', 'Pierwszy.'),
        (2, 'Second.', 'Drugi.'),
    ])
    excel = _make_excel_for_merge(['First. Second.'])
    result = merge_xliff22_to_excel(xliff, excel)
    assert result['rows_written'] == 1
    wb = load_workbook(str(excel))
    ws = wb.active
    assert ws.cell(row=2, column=2).value == 'Pierwszy. Drugi.'


def test_merger_skips_empty_targets():
    xliff = _make_xliff([
        (2, 'Hello.', ''),
        (3, 'World.', 'Świat.'),
    ])
    excel = _make_excel_for_merge(['Hello.', 'World.'])
    result = merge_xliff22_to_excel(xliff, excel)
    assert result['rows_written'] == 1
    wb = load_workbook(str(excel))
    ws = wb.active
    assert ws.cell(row=2, column=2).value is None
    assert ws.cell(row=3, column=2).value == 'Świat.'


def test_merger_raises_on_missing_metadata():
    root = etree.Element(f'{{{NS22}}}xliff', nsmap={None: NS22})
    file_elem = etree.SubElement(root, f'{{{NS22}}}file')
    file_elem.set('id', 'test.xlsx')
    tmp = tempfile.NamedTemporaryFile(suffix='.xliff', delete=False)
    etree.ElementTree(root).write(tmp.name)
    tmp.close()
    _TEMP_FILES.append(tmp.name)
    with pytest.raises(ValueError, match='x-excel-tgt-col'):
        merge_xliff22_to_excel(tmp.name, '/tmp/irrelevant.xlsx')


def test_merger_row_beyond_sheet_skipped():
    xliff = _make_xliff([
        (2, 'Present.', 'Obecny.'),
        (999, 'Beyond.', 'Poza.'),  # row 999, sheet only has row 2
    ])
    excel = _make_excel_for_merge(['Present.'])
    result = merge_xliff22_to_excel(xliff, excel)
    assert result['rows_written'] == 1
